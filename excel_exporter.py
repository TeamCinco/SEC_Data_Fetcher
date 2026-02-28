import pandas as pd
import re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Formats parsed XBRL data into styled Excel sheets.
    Model scaffolding (Cover, Assumptions, DCF) lives in excel_generator."""

    PROJECTION_YEARS = 5

    HEADER_FILL = PatternFill("solid", fgColor="E7EEF7")
    YEAR_LABEL_FILL = PatternFill("solid", fgColor="D6E4F0")
    ZEBRA_FILL = PatternFill("solid", fgColor="F7F9FC")
    SUBTOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
    PROJ_HEADER_FILL = PatternFill("solid", fgColor="D6E4F0")

    CURRENCY_FMT = '#,##0;(#,##0);"-"'
    THIN_BORDER_BTM = Border(bottom=Side("thin", color="B0B0B0"))
    THIN_BORDER_TOP = Border(top=Side("thin", color="808080"))

    TOTAL_KEYWORDS = {
        "total", "net income", "net loss", "gross profit",
        "operating income", "ebit", "ebitda",
        "income before", "loss before", "net revenue",
    }

    # ── Public ──

    def write_data_sheets(self, financials, writer):
        """Write All Facts + statement sheets into an open pd.ExcelWriter."""
        facts = financials["facts"]
        statements = financials["statements"]

        pivot = facts.pivot_table(index="concept", columns="date", values="value", aggfunc="first")
        pivot = self._sort_cols(pivot)
        pivot.to_excel(writer, sheet_name="All Facts")
        self._beautify(writer.book["All Facts"], year_labels=True, projections=False)

        for name, df in statements.items():
            clean = df.copy()
            if "level" in clean.columns and "label" in clean.columns:
                clean["label"] = clean.apply(
                    lambda r: "    " * int(r["level"]) + str(r["label"]), axis=1
                )
            clean = self._sort_df_cols(clean)
            clean.to_excel(writer, sheet_name=name)
            self._beautify(writer.book[name], year_labels=True, projections=True)

    def count_date_columns(self, statements):
        for _, df in statements.items():
            n = sum(1 for c in df.columns if self._is_date(c))
            if n > 0:
                return n
        return 3

    # ── Column helpers ──

    def _sort_key(self, col):
        s = str(col)
        m = re.match(r'[Yy]ear\s*([+-]?\d+)', s)
        if m:
            return (0, int(m.group(1)), s)
        try:
            return (1, pd.Timestamp(s).timestamp(), s)
        except (ValueError, TypeError):
            pass
        if re.match(r'^\d{4}$', s):
            return (1, int(s), s)
        return (2, 0, s)

    def _is_date(self, col):
        s = str(col)
        if re.match(r'[Yy]ear\s*[+-]?\d+', s):
            return True
        try:
            pd.Timestamp(s)
            return True
        except (ValueError, TypeError):
            return bool(re.match(r'^\d{4}$', s))

    def _sort_cols(self, pivot):
        return pivot[sorted(pivot.columns, key=self._sort_key)]

    def _sort_df_cols(self, df):
        meta = [c for c in df.columns if not self._is_date(c)]
        dates = sorted([c for c in df.columns if self._is_date(c)], key=self._sort_key)
        return df[meta + dates]

    # ── Formatting engine ──

    def _beautify(self, ws, year_labels=False, projections=False):
        hdr_font = Font(bold=True, size=12)
        data_font = Font(size=11)
        bold_font = Font(bold=True, size=11)
        yr_font = Font(size=10, bold=True, color="4472C4")
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        max_row, max_col = ws.max_row, ws.max_column

        date_hdrs = [
            (c, str(ws.cell(row=1, column=c).value or ""))
            for c in range(1, max_col + 1)
            if self._is_date(str(ws.cell(row=1, column=c).value or ""))
        ]
        num_hist = len(date_hdrs)

        # Year label row (inserted above headers)
        if year_labels and date_hdrs:
            ws.insert_rows(1)
            max_row += 1
            for i, (col, _) in enumerate(date_hdrs):
                cell = ws.cell(row=1, column=col, value=f"Year {i - (num_hist - 1)}")
                cell.font, cell.fill, cell.alignment = yr_font, self.YEAR_LABEL_FILL, center
            hdr_row, data_row = 2, 3
        else:
            hdr_row, data_row = 1, 2

        # Header styling
        for c in range(1, max_col + 1):
            cell = ws.cell(row=hdr_row, column=c)
            cell.font, cell.fill, cell.alignment = hdr_font, self.HEADER_FILL, center

        # Data rows
        for r in range(data_row, max_row + 1):
            lbl = str(ws.cell(row=r, column=1).value or "").strip().lower()
            total = any(kw in lbl for kw in self.TOTAL_KEYWORDS)

            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = bold_font if total else data_font
                if total:
                    cell.border = self.THIN_BORDER_TOP
                cell.alignment = left if c == 1 else center
                if c > 1:
                    cell.number_format = self.CURRENCY_FMT

            fill = self.SUBTOTAL_FILL if total else (self.ZEBRA_FILL if r % 2 == 0 else None)
            if fill:
                for c in range(1, max_col + 1):
                    ws.cell(row=r, column=c).fill = fill

        # Projection columns
        if projections and date_hdrs:
            last_yr = None
            for _, h in reversed(date_hdrs):
                m = re.search(r'(\d{4})', h)
                if m:
                    last_yr = int(m.group(1))
                    break
            if last_yr:
                for i in range(1, self.PROJECTION_YEARS + 1):
                    pc = max_col + i
                    if year_labels:
                        cell = ws.cell(row=1, column=pc, value=f"Year {i}")
                        cell.font, cell.fill, cell.alignment = yr_font, self.YEAR_LABEL_FILL, center
                    cell = ws.cell(row=hdr_row, column=pc, value=f"{last_yr + i}E")
                    cell.font = Font(bold=True, size=12, color="1F4E79")
                    cell.fill, cell.alignment = self.PROJ_HEADER_FILL, center
                    for r in range(data_row, max_row + 1):
                        cell = ws.cell(row=r, column=pc)
                        cell.number_format = self.CURRENCY_FMT
                        cell.alignment = center
                        cell.border = self.THIN_BORDER_BTM
                    ws.column_dimensions[get_column_letter(pc)].width = 18

        # Auto-width + freeze
        for c in range(1, ws.max_column + 1):
            mx = max((len(str(ws.cell(row=r, column=c).value or ""))
                       for r in range(1, ws.max_row + 1)), default=8)
            ws.column_dimensions[get_column_letter(c)].width = min(mx + 4, 50)

        ws.freeze_panes = "B3" if year_labels else "B2"