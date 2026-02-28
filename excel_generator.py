import pandas as pd
import re
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from xbrl_parser import SECXBRLParser
from excel_exporter import ExcelExporter


# ── Shared styles ──

SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
HEADER_FILL = PatternFill("solid", fgColor="E7EEF7")
YEAR_LABEL_FILL = PatternFill("solid", fgColor="D6E4F0")
PROJ_HEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBTOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")

CURRENCY_FMT = '$#,##0;($#,##0);"-"'
PLAIN_FMT = '#,##0;(#,##0);"-"'
PCT_FMT = '0.0%'
MULT_FMT = '0.0"x"'

BORDER_TOP = Border(top=Side("thin", color="808080"))

SECTION_FONT = Font(size=12, bold=True, color="1F4E79")
LABEL_FONT = Font(size=11, color="333333")
BOLD_LABEL = Font(size=11, bold=True, color="333333")
BODY_FONT = Font(size=11, color="333333")
INPUT_FONT = Font(size=11, color="0000FF", bold=True)
NOTE_FONT = Font(size=10, italic=True, color="808080")
HEADER_FONT = Font(size=11, bold=True, color="1F4E79")
YR_FONT = Font(size=10, bold=True, color="4472C4")
CENTER = Alignment(horizontal="center", vertical="center")


# ═══════════════════════════════════════════════════════════════
#  Main entry point
# ═══════════════════════════════════════════════════════════════

def generate_excel_from_filing(filing_url, output_path=None):
    parser = SECXBRLParser()
    exporter = ExcelExporter()

    financials = parser.extract_statements(filing_url)
    statements = financials["statements"]
    num_hist = exporter.count_date_columns(statements)

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # Placeholder sheets so they exist before data sheets
        pd.DataFrame().to_excel(writer, sheet_name="Cover")
        pd.DataFrame().to_excel(writer, sheet_name="Assumptions")

        # Data sheets (All Facts + statements)
        exporter.write_data_sheets(financials, writer)

        # DCF sheet
        pd.DataFrame().to_excel(writer, sheet_name="DCF Valuation")

        wb = writer.book
        build_cover(wb["Cover"])
        build_assumptions(wb["Assumptions"], exporter.PROJECTION_YEARS)
        build_dcf(wb["DCF Valuation"], num_hist, exporter.PROJECTION_YEARS)

        # Reorder: Cover, Assumptions, data..., DCF last
        order = ["Cover", "Assumptions"]
        order += [s for s in wb.sheetnames if s not in order and s != "DCF Valuation"]
        order += ["DCF Valuation"]
        wb._sheets = [wb[s] for s in order if s in wb.sheetnames]

    output.seek(0)

    if output_path:
        with open(output_path, "wb") as f:
            f.write(output.getvalue())
        return True

    return output


# ═══════════════════════════════════════════════════════════════
#  Cover Sheet
# ═══════════════════════════════════════════════════════════════

def build_cover(ws):
    ws.sheet_properties.tabColor = "1F4E79"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 40

    ws.merge_cells("B3:C3")
    ws["B3"] = "DCF Valuation Model"
    ws["B3"].font = Font(size=22, bold=True, color="1F4E79")

    ws.merge_cells("B4:C4")
    ws["B4"] = "Discounted Cash Flow Analysis"
    ws["B4"].font = Font(size=14, color="4472C4")

    # Company info
    _section(ws, "B6:C6", "Company Information")
    for i, (lbl, val) in enumerate([
        ("Company Name:", ""), ("Ticker:", ""),
        ("Filing Source:", "SEC EDGAR XBRL"),
        ("Generated:", datetime.now().strftime("%B %d, %Y")),
        ("Analyst:", ""),
    ]):
        r = 7 + i
        ws[f"B{r}"], ws[f"B{r}"].font = lbl, BODY_FONT
        ws[f"C{r}"] = val
        ws[f"C{r}"].font = INPUT_FONT if val == "" else BODY_FONT

    # Sheet guide
    _section(ws, "B13:C13", "Sheet Guide")
    for i, (sheet, desc) in enumerate([
        ("Cover", "Company info & model overview"),
        ("Assumptions", "WACC, growth rates, margins, terminal value"),
        ("All Facts", "Raw XBRL data pivot (oldest → newest)"),
        ("Statement Sheets", "Historical + projection columns"),
        ("DCF Valuation", "Blank DCF template for your formulas"),
    ]):
        r = 14 + i
        ws[f"B{r}"] = sheet
        ws[f"B{r}"].font = Font(size=11, bold=True, color="333333")
        ws[f"C{r}"], ws[f"C{r}"].font = desc, BODY_FONT

    # Year label legend
    _section(ws, "B20:C20", "Year Label Convention")
    for i, (lbl, desc) in enumerate([
        ("Year -N", "Historical periods (oldest data)"),
        ("Year 0", "Most recent reported period (present)"),
        ("Year 1 … Year 5", "Projection periods (your forecast)"),
    ]):
        r = 21 + i
        ws[f"B{r}"], ws[f"B{r}"].font = lbl, YR_FONT
        ws[f"C{r}"], ws[f"C{r}"].font = desc, BODY_FONT

    # Color coding
    _section(ws, "B25:C25", "Color Coding")
    for i, (lbl, desc, clr, fill) in enumerate([
        ("Blue text", "Hardcoded inputs / assumptions", "0000FF", None),
        ("Black text", "Formulas and calculations", "000000", None),
        ("Green text", "Cross-sheet references", "008000", None),
        ("Yellow fill", "Cells needing your input", None, INPUT_FILL),
    ]):
        r = 26 + i
        ws[f"B{r}"] = lbl
        ws[f"B{r}"].font = Font(size=11, bold=True, color=clr or "000000")
        if fill:
            ws[f"B{r}"].fill = fill
        ws[f"C{r}"], ws[f"C{r}"].font = desc, BODY_FONT


# ═══════════════════════════════════════════════════════════════
#  Assumptions Sheet
# ═══════════════════════════════════════════════════════════════

def build_assumptions(ws, proj_years=5):
    ws.sheet_properties.tabColor = "FFC000"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 45

    row = _assumption_block(ws, 2, "Weighted Average Cost of Capital (WACC)", [
        ("Risk-Free Rate",       PCT_FMT,  "10-Year Treasury yield"),
        ("Equity Risk Premium",  PCT_FMT,  "Historical market premium"),
        ("Beta (Levered)",       "0.00",   "Regression or comparable"),
        ("Cost of Equity (CAPM)",PCT_FMT,  "= Rf + Beta x ERP"),
        ("Pre-Tax Cost of Debt", PCT_FMT,  "Interest expense / total debt"),
        ("Tax Rate",             PCT_FMT,  "Effective or marginal"),
        ("After-Tax Cost of Debt", PCT_FMT,"= Kd x (1 - Tax Rate)"),
        ("Debt / Total Capital", PCT_FMT,  "Market value weights"),
        ("Equity / Total Capital", PCT_FMT,"Market value weights"),
        ("WACC",                 PCT_FMT,  "= Ke x (E/V) + Kd(1-t) x (D/V)"),
    ])

    # Growth & Margins — one input cell per projection year
    row = _growth_block(ws, row + 1, proj_years, "Growth & Margin Assumptions", [
        "Revenue Growth Rate",
        "Gross Margin",
        "Operating Margin",
        "Net Margin",
        "CapEx as % of Revenue",
        "D&A as % of Revenue",
        "Change in NWC as % of Revenue",
    ])

    row = _assumption_block(ws, row + 1, "Terminal Value Assumptions", [
        ("Terminal Growth Rate (g)",  PCT_FMT,  "Long-term GDP growth proxy"),
        ("Exit Multiple (EV/EBITDA)", MULT_FMT, "Comparable company multiple"),
        ("Terminal Value Method",     "@",      "Gordon Growth or Exit Multiple"),
    ])

    _assumption_block(ws, row + 1, "Equity Bridge", [
        ("Shares Outstanding (diluted)", '#,##0',       "From latest filing"),
        ("Net Debt (Debt - Cash)",       CURRENCY_FMT,  "Balance sheet"),
        ("Minority Interest",            CURRENCY_FMT,  "If applicable"),
        ("Preferred Equity",             CURRENCY_FMT,  "If applicable"),
    ])

    ws.freeze_panes = "B1"


# ═══════════════════════════════════════════════════════════════
#  DCF Valuation Sheet
# ═══════════════════════════════════════════════════════════════

def build_dcf(ws, num_hist=3, proj_years=5):
    ws.sheet_properties.tabColor = "00B050"
    total = num_hist + proj_years
    cs = 3  # data starts at column C

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    for i in range(total):
        ws.column_dimensions[get_column_letter(cs + i)].width = 18

    end_col = get_column_letter(cs + total - 1)

    # Row 1: Year labels
    ws["B1"], ws["B1"].font = "Year Label", YR_FONT
    for i in range(total):
        cell = ws.cell(row=1, column=cs + i, value=f"Year {i - (num_hist - 1)}")
        cell.font, cell.fill, cell.alignment = YR_FONT, YEAR_LABEL_FILL, CENTER

    # Row 2: Period placeholders
    ws["B2"], ws["B2"].font = "Period", HEADER_FONT
    for i in range(total):
        cell = ws.cell(row=2, column=cs + i, value="")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL if i < num_hist else PROJ_HEADER_FILL
        cell.alignment = CENTER

    # FCF build-up
    r = 4
    _dcf_section(ws, r, end_col, "Free Cash Flow Build-Up")
    fcf = [
        ("Revenue", False), ("Revenue Growth %", False), ("", False),
        ("EBIT (Operating Income)", False), ("EBIT Margin %", False), ("", False),
        ("Tax Rate", False), ("NOPAT", True), ("", False),
        ("(+) Depreciation & Amortization", False),
        ("(-) Capital Expenditures", False),
        ("(-) Change in Net Working Capital", False), ("", False),
        ("Unlevered Free Cash Flow (UFCF)", True),
    ]
    _dcf_rows(ws, r + 1, fcf, cs, total)

    # Discounting
    dr = r + 1 + len(fcf) + 1
    _dcf_section(ws, dr, end_col, "Discounting")
    disc = [
        ("WACC", False), ("Discount Period", False),
        ("Discount Factor", False), ("PV of UFCF", True),
    ]
    _dcf_rows(ws, dr + 1, disc, cs, total)

    # Terminal value
    tr = dr + 1 + len(disc) + 1
    _dcf_section(ws, tr, end_col, "Terminal Value")
    for i, lbl in enumerate([
        "Terminal Growth Rate (g)", "WACC", "Terminal Year FCF",
        "Terminal Value (Gordon Growth)", "Terminal Value (Exit Multiple)",
        "", "Selected Terminal Value", "PV of Terminal Value",
    ]):
        row = tr + 1 + i
        ws[f"B{row}"] = lbl
        ws[f"B{row}"].font = BOLD_LABEL if "Selected" in lbl or "PV of" in lbl else LABEL_FONT

    # Enterprise & equity value
    er = tr + 10
    _dcf_section(ws, er, end_col, "Enterprise & Equity Value")
    ev_items = [
        ("Sum of PV of FCFs", True), ("(+) PV of Terminal Value", False),
        ("Enterprise Value", True), ("", False),
        ("(-) Total Debt", False), ("(+) Cash & Equivalents", False),
        ("(-) Minority Interest", False), ("(-) Preferred Equity", False),
        ("", False), ("Equity Value", True),
        ("Shares Outstanding (diluted)", False), ("", False),
        ("Implied Share Price", True), ("Current Market Price", False),
        ("Upside / (Downside)", True),
    ]
    for i, (lbl, bold) in enumerate(ev_items):
        row = er + 1 + i
        ws[f"B{row}"] = lbl
        ws[f"B{row}"].font = BOLD_LABEL if bold else LABEL_FONT
        if bold and lbl:
            c = ws.cell(row=row, column=cs)
            c.fill, c.border = SUBTOTAL_FILL, BORDER_TOP

    # Sensitivity
    sr = er + 1 + len(ev_items) + 1
    _dcf_section(ws, sr, end_col, "Sensitivity Analysis (WACC vs. Terminal Growth)")
    ws[f"B{sr + 1}"] = "(Build your data table here)"
    ws[f"B{sr + 1}"].font = NOTE_FONT

    ws.freeze_panes = "C3"


# ═══════════════════════════════════════════════════════════════
#  Helpers
# ═══════════════════════════════════════════════════════════════

def _section(ws, merge_range, title):
    ws.merge_cells(merge_range)
    top_left = merge_range.split(":")[0]
    ws[top_left] = title
    ws[top_left].font = SECTION_FONT
    ws[top_left].fill = SECTION_FILL


def _assumption_block(ws, start, title, items):
    ws[f"B{start}"] = title
    ws[f"B{start}"].font = SECTION_FONT
    ws[f"B{start}"].fill = SECTION_FILL
    ws.merge_cells(f"B{start}:D{start}")
    for i, (lbl, fmt, note) in enumerate(items):
        r = start + 1 + i
        ws[f"B{r}"], ws[f"B{r}"].font = lbl, LABEL_FONT
        ws[f"C{r}"].font = INPUT_FONT
        ws[f"C{r}"].number_format = fmt
        ws[f"C{r}"].fill = INPUT_FILL
        ws[f"D{r}"], ws[f"D{r}"].font = note, NOTE_FONT
    return start + 1 + len(items)


def _growth_block(ws, start, proj_years, title, labels):
    """Growth & margin section with one input column per projection year."""
    last_col_letter = get_column_letter(2 + proj_years)  # B + N year columns

    # Section header
    ws[f"B{start}"] = title
    ws[f"B{start}"].font = SECTION_FONT
    ws[f"B{start}"].fill = SECTION_FILL
    ws.merge_cells(f"B{start}:{last_col_letter}{start}")

    # Year column headers (Year 1, Year 2, ...)
    for y in range(1, proj_years + 1):
        col = 2 + y  # C=3, D=4, ...
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 16
        cell = ws.cell(row=start + 1, column=col, value=f"Year {y}")
        cell.font = YR_FONT
        cell.fill = YEAR_LABEL_FILL
        cell.alignment = CENTER

    ws[f"B{start + 1}"] = ""
    ws[f"B{start + 1}"].font = LABEL_FONT

    # One row per metric, one yellow input cell per year
    for i, lbl in enumerate(labels):
        r = start + 2 + i
        ws[f"B{r}"] = lbl
        ws[f"B{r}"].font = LABEL_FONT
        for y in range(1, proj_years + 1):
            cell = ws.cell(row=r, column=2 + y)
            cell.font = INPUT_FONT
            cell.number_format = PCT_FMT
            cell.fill = INPUT_FILL
            cell.alignment = CENTER

    return start + 2 + len(labels)


def _dcf_section(ws, row, end_col, title):
    ws[f"B{row}"] = title
    ws[f"B{row}"].font = SECTION_FONT
    ws[f"B{row}"].fill = SECTION_FILL
    ws.merge_cells(f"B{row}:{end_col}{row}")


def _dcf_rows(ws, start, items, col_start, total_cols):
    for i, (lbl, is_total) in enumerate(items):
        r = start + i
        ws[f"B{r}"] = lbl
        ws[f"B{r}"].font = BOLD_LABEL if is_total else LABEL_FONT
        for c in range(col_start, col_start + total_cols):
            cell = ws.cell(row=r, column=c)
            cell.number_format = PLAIN_FMT
            cell.alignment = CENTER
            if is_total:
                cell.fill = SUBTOTAL_FILL
                cell.border = BORDER_TOP