import re
import requests
import tempfile
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class BulletProofConsolidatedParser:

    TARGET_HEADINGS = [
        "consolidated balance sheets",
        "consolidated statements of operations",
        "consolidated statements of comprehensive",
        "consolidated statements of cash flows",
        "consolidated statements of shareholders",
        "consolidated statements of stockholders"
    ]

    def __init__(self, html_file, output_file):
        self.html_file = Path(html_file)
        self.output_file = Path(output_file)
        self.soup = None
        self.context_map = {}
        self.valid_year_ends = set()

    # ---------------------------------------------------------
    # LOAD AS XML (CRITICAL)
    # ---------------------------------------------------------

    def load(self):
        with open(self.html_file, "r", encoding="utf-8", errors="ignore") as f:
            self.soup = BeautifulSoup(f.read(), "xml")

    # ---------------------------------------------------------
    # BUILD CONTEXT MAP (NO SEGMENTS)
    # ---------------------------------------------------------

    def build_context_map(self):

        for ctx in self.soup.find_all("context"):

            ctx_id = ctx.get("id")
            if not ctx_id:
                continue

            # Skip dimensional contexts
            if ctx.find("segment"):
                continue

            period = ctx.find("period")
            if not period:
                continue

            instant = period.find("instant")
            start = period.find("startDate")
            end = period.find("endDate")

            data = {"type": None, "start": None, "end": None, "instant": None}

            if instant:
                data["type"] = "instant"
                data["instant"] = instant.text.strip()

            elif start and end:
                data["type"] = "duration"
                data["start"] = start.text.strip()
                data["end"] = end.text.strip()

            self.context_map[ctx_id] = data

    # ---------------------------------------------------------
    # DETECT VALID FISCAL YEAR-END DATES FROM HEADER
    # ---------------------------------------------------------

    def detect_year_ends_from_table(self, table):

        dates = set()

        header_text = table.get_text(" ", strip=True)

        # Look for June 27, 2025 etc
        matches = re.findall(r"[A-Za-z]+\s+\d{1,2},\s+\d{4}", header_text)

        for match in matches:
            try:
                dt = datetime.strptime(match, "%B %d, %Y")
                dates.add(dt.strftime("%Y-%m-%d"))
            except:
                continue

        return dates

    # ---------------------------------------------------------
    # FILTER FULL YEAR DURATION
    # ---------------------------------------------------------

    def is_full_year(self, context):

        if context["type"] != "duration":
            return False

        try:
            d1 = datetime.fromisoformat(context["start"])
            d2 = datetime.fromisoformat(context["end"])
            return (d2 - d1).days > 350
        except:
            return False

    # ---------------------------------------------------------
    # FIND CONSOLIDATED TABLES
    # ---------------------------------------------------------

    def find_statement_tables(self):

        tables = []

        headings = self.soup.find_all(
            lambda tag: tag.name in ["p", "div", "strong", "h1", "h2", "h3"]
            and tag.get_text(strip=True)
        )

        for heading in headings:

            text = heading.get_text(strip=True).lower()

            if any(t in text for t in self.TARGET_HEADINGS):

                table = heading.find_next("table")

                if table:
                    tables.append((heading.get_text(strip=True), table))

        return tables

    # ---------------------------------------------------------
    # CLEAN NUMERIC TAG
    # ---------------------------------------------------------

    def parse_numeric(self, tag):

        if tag.get("format") == "ixt:fixed-zero":
            return 0.0

        text = tag.get_text(strip=True)
        text = text.replace(",", "").replace("$", "").strip()

        if text in ["—", "-", ""]:
            return None

        try:
            value = float(text)
        except:
            return None

        scale = int(tag.get("scale", "0"))
        value *= 10 ** scale

        if tag.get("sign") == "-":
            value = -abs(value)

        return value

    # ---------------------------------------------------------
    # PARSE TABLE WITH STRICT FILTERING
    # ---------------------------------------------------------

    def parse_table(self, table):

        data = defaultdict(dict)

        # detect valid fiscal year-ends
        self.valid_year_ends = self.detect_year_ends_from_table(table)

        for row in table.find_all("tr"):

            cells = row.find_all(["td", "th"])
            if not cells:
                continue

            label = None

            # Extract first meaningful text as label
            for cell in cells:
                text = cell.get_text(strip=True)
                if text and text not in ["$", "—", "-"]:
                    label = text
                    break

            if not label:
                continue

            numeric_tags = row.find_all("ix:nonFraction")

            if not numeric_tags:
                continue

            for tag in numeric_tags:

                context_id = tag.get("contextRef")
                context = self.context_map.get(context_id)

                if not context:
                    continue

                # Determine date
                if context["type"] == "instant":
                    date = context["instant"]
                else:
                    if not self.is_full_year(context):
                        continue
                    date = context["end"]

                # Only keep dates that appear in header
                if self.valid_year_ends and date not in self.valid_year_ends:
                    continue

                value = self.parse_numeric(tag)
                if value is None:
                    continue

                data[label][date] = value

        return data

    # ---------------------------------------------------------
    # WRITE TO EXCEL
    # ---------------------------------------------------------

    def write_excel(self):

        wb = Workbook()
        wb.remove(wb.active)

        used_names = set()

        for title, table in self.find_statement_tables():

            parsed = self.parse_table(table)

            if not parsed:
                continue

            sheet_name = title[:31]
            base = sheet_name
            counter = 1

            while sheet_name in used_names:
                sheet_name = f"{base[:28]}_{counter}"
                counter += 1

            used_names.add(sheet_name)

            ws = wb.create_sheet(sheet_name)

            dates = sorted(
                {d for row in parsed.values() for d in row.keys()}
            )

            ws.cell(row=1, column=1, value="Line Item").font = Font(bold=True)

            for col, date in enumerate(dates, start=2):
                ws.cell(row=1, column=col, value=date).font = Font(bold=True)

            for r_idx, (label, values) in enumerate(parsed.items(), start=2):
                ws.cell(row=r_idx, column=1, value=label)

                for c_idx, date in enumerate(dates, start=2):
                    if date in values:
                        ws.cell(
                            row=r_idx,
                            column=c_idx,
                            value=values[date]
                        ).alignment = Alignment(horizontal="right")

            ws.column_dimensions["A"].width = 55
            for col in range(2, len(dates) + 2):
                ws.column_dimensions[get_column_letter(col)].width = 18

        wb.save(self.output_file)

    # ---------------------------------------------------------
    # RUN
    # ---------------------------------------------------------

    def run(self):
        self.load()
        self.build_context_map()
        self.write_excel()


def parse_xbrl_to_excel(html_url, output_path):

    headers = {"User-Agent": "Research App contact@example.com"}

    response = requests.get(html_url, headers=headers, timeout=30)
    response.raise_for_status()

    with tempfile.NamedTemporaryFile(delete=False, suffix=".html") as tmp:
        tmp.write(response.content)
        tmp_path = tmp.name

    parser = BulletProofConsolidatedParser(tmp_path, output_path)
    parser.run()

    Path(tmp_path).unlink()

    return output_path
